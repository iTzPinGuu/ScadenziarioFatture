#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Scadenziario Commerciale Premium — VIDALOCA di Michela Vidale
Versione 6.8 - Dashboard Avanzata, Bugfix Rubrica, Filtri e Totali Dinamici
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import csv
import os
import webbrowser
import urllib.request
import io
from datetime import datetime, date, timedelta
import calendar
import re

try:
    from PIL import Image, ImageTk
    PILLOW_OK = True
except ImportError:
    PILLOW_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

DB_FILE = r"O:\GIOVANNI PIO\DATABASE PER VIDALOCA\scadenziario.db"
LOGO_URL = "https://www.truccoloangelo.com/wp-content/uploads/2026/05/Logo-Vidaloca.png"

C = {
    "bg":          "#fdfbf7",  
    "surface":     "#ffffff",  
    "border":      "#e7e1d5",  
    "text":        "#3d2d1f",  
    "muted":       "#857364",  
    "accent":      "#c5a059",  
    "accent_dark": "#8c6f34",  
    "teal_accent": "#1e6b7b",  
    "scaduta":     "#fdf2f2",  
    "scaduta_fg":  "#991b1b",  
    "urgente":     "#fef3c7",  
    "urgente_fg":  "#92400e",  
    "ok":          "#f0fdf4",  
    "ok_fg":       "#166534",  
    "red":         "#ef4444",  
    "green":       "#10b981",  
    "aperti":      "#eff6ff",  # Aggiunto per nuovo filtro
    "aperti_fg":   "#1e3a8a",  # Aggiunto per nuovo filtro
}

STATI_PASSIVO   = ["Da pagare", "Parziale", "Pagata", "Insoluta"]
STATI_ATTIVO    = ["Da incassare", "Parziale", "Incassata", "Insoluta"]

def get_conn():
    return sqlite3.connect(DB_FILE)

def init_db():
    with get_conn() as conn:
        c = conn.cursor()
        c.execute("""
            CREATE TABLE IF NOT EXISTS aziende (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT UNIQUE NOT NULL,
                condizioni TEXT NOT NULL DEFAULT '30 Giorni'
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS config_condizioni (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT UNIQUE NOT NULL
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS config_metodi_pag (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT UNIQUE NOT NULL
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS config_tipi_doc (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT UNIQUE NOT NULL
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS fatture (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                tipo        TEXT NOT NULL DEFAULT 'passivo',
                numero      TEXT NOT NULL,
                anagrafica  TEXT NOT NULL,
                data_doc    TEXT,
                data_scad   TEXT,
                data_pag    TEXT,
                importo     REAL NOT NULL,
                pagato      REAL NOT NULL DEFAULT 0,
                stato       TEXT NOT NULL DEFAULT 'Da pagare',
                metodo      TEXT,
                tipo_doc    TEXT DEFAULT 'Fattura',
                note        TEXT,
                data_ins    TEXT DEFAULT (date('now'))
            )
        """)
        
        # Aggiornamento schema per database esistenti senza la colonna data_pag
        try:
            c.execute("ALTER TABLE fatture ADD COLUMN data_pag TEXT")
        except Exception:
            pass
            
        c.execute("""
            CREATE TABLE IF NOT EXISTS pagamenti (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                fattura_id  INTEGER NOT NULL,
                data_pag    TEXT NOT NULL,
                importo     REAL NOT NULL,
                metodo      TEXT,
                note        TEXT,
                FOREIGN KEY(fattura_id) REFERENCES fatture(id)
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS corrispettivi (
                anno INTEGER NOT NULL,
                mese INTEGER NOT NULL,
                giorno INTEGER NOT NULL,
                totale REAL DEFAULT 0.0,
                iva_1 REAL DEFAULT 0.0,
                iva_2 REAL DEFAULT 0.0,
                iva_3 REAL DEFAULT 0.0,
                iva_4 REAL DEFAULT 0.0,
                ventilati REAL DEFAULT 0.0,
                esenti REAL DEFAULT 0.0,
                autoconsumo REAL DEFAULT 0.0,
                note TEXT DEFAULT '',
                PRIMARY KEY (anno, mese, giorno)
            )
        """)
        c.execute("CREATE INDEX IF NOT EXISTS idx_fatture_ricerca ON fatture (numero, anagrafica, tipo)")
        
        c.execute("SELECT COUNT(*) FROM config_condizioni")
        if c.fetchone()[0] == 0:
            for cond in ["30 Giorni", "15 Giorni", "60 Giorni", "90 Giorni", "Vista fattura", "Contanti", "Riba 30 gg", "Riba 60 gg"]:
                c.execute("INSERT OR IGNORE INTO config_condizioni (nome) VALUES (?)", (cond,))
                
        c.execute("SELECT COUNT(*) FROM config_metodi_pag")
        if c.fetchone()[0] == 0:
            for met in ["Bonifico", "RID/SDD", "Contanti", "Assegno", "Carta", "Riba", "Altro"]:
                c.execute("INSERT OR IGNORE INTO config_metodi_pag (nome) VALUES (?)", (met,))

        c.execute("SELECT COUNT(*) FROM config_tipi_doc")
        if c.fetchone()[0] == 0:
            for td in ["Fattura", "Nota di credito", "Parcella", "Fattura differita", "Altro"]:
                c.execute("INSERT OR IGNORE INTO config_tipi_doc (nome) VALUES (?)", (td,))

def get_condizioni_pagamento():
    try:
        with get_conn() as conn:
            rows = conn.execute("SELECT nome FROM config_condizioni ORDER BY id ASC").fetchall()
            if rows: return [r[0] for r in rows]
    except Exception: pass
    return ["30 Giorni", "15 Giorni", "60 Giorni", "90 Giorni", "Vista fattura", "Contanti", "Riba 30 gg", "Riba 60 gg"]

def get_metodi_pagamento():
    try:
        with get_conn() as conn:
            rows = conn.execute("SELECT nome FROM config_metodi_pag ORDER BY id ASC").fetchall()
            if rows: return [r[0] for r in rows]
    except Exception: pass
    return ["Bonifico", "Contanti", "Carta", "Altro"]

def get_tipi_documento():
    try:
        with get_conn() as conn:
            rows = conn.execute("SELECT nome FROM config_tipi_doc ORDER BY id ASC").fetchall()
            if rows: return [r[0] for r in rows]
    except Exception: pass
    return ["Fattura", "Nota di credito", "Altro"]

def parse_date(val):
    if not val: return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d") if hasattr(val, "hour") else val.isoformat()
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try: return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError: continue
    return s

def parse_importo(val):
    if val is None or val == "": return 0.0
    s = str(val).strip().replace(" ", "")
    if "," in s and "." in s: s = s.replace(".", "").replace(",", ".")
    else: s = s.replace(",", ".")
    try: return float(s)
    except ValueError: return 0.0

def fmt_date(iso):
    if not iso: return ""
    try: return datetime.strptime(iso, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception: return iso

def fmt_num(v):
    try: return f"{float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception: return str(v)

def calcola_scadenza_automatico(c_db, nome_azienda, data_doc_iso):
    if not data_doc_iso: return ""
    res = c_db.execute("SELECT condizioni FROM aziende WHERE nome = ?", (nome_azienda,)).fetchone()
    cond = res[0] if res else "30 Giorni"
    giorni = 30
    if "vista fattura" in cond.lower() or "contanti" in cond.lower(): giorni = 0
    else:
        match = re.search(r'\d+', cond)
        if match: giorni = int(match.group())
    try:
        dt = datetime.strptime(data_doc_iso, "%Y-%m-%d").date()
        return (dt + timedelta(days=giorni)).strftime("%Y-%m-%d")
    except Exception: return ""

def mappa_stato_fte(stato_fte, tipo):
    if not stato_fte: return "Da pagare" if tipo == "passivo" else "Da incassare"
    s = stato_fte.strip().lower()
    if s in ("pagato", "pagata", "incassata", "incassato"): return "Pagata" if tipo == "passivo" else "Incassata"
    if s == "parziale": return "Parziale"
    return "Da pagare" if tipo == "passivo" else "Da incassare"


# ─── FINESTRE DI ACQUISIZIONE E DIALOGHI MULTIPLI ────────────────────────────
class DateEntry(tk.Frame):
    def __init__(self, master, **kw):
        super().__init__(master, bg=C["surface"])
        self._var = tk.StringVar()
        self._entry = tk.Entry(self, textvariable=self._var, width=12, font=("Segoe UI", 10), relief="solid", bd=1, highlightthickness=0, borderwidth=1, fg=C["text"])
        self._entry.pack(side="left", ipady=3)
        btn = tk.Button(self, text="📅", command=self._open_cal, relief="flat", bg=C["surface"], fg=C["accent"], cursor="hand2", font=("Segoe UI", 10))
        btn.pack(side="left", padx=(4, 0))
        self._entry.bind("<FocusOut>", self._on_focus_out)

    def _on_focus_out(self, e=None):
        v = self._var.get().strip()
        if not v: return
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                d = datetime.strptime(v, fmt)
                self._var.set(d.strftime("%d/%m/%Y"))
                return
            except ValueError: continue

    def _open_cal(self):
        top = tk.Toplevel(self)
        top.title("Seleziona Data")
        top.resizable(False, False)
        top.geometry(f"+{self._entry.winfo_rootx()}+{self._entry.winfo_rooty() + 30}")
        CalPicker(top, self._var, top).pack()
        top.grab_set()

    def get(self):
        v = self._var.get().strip()
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try: return datetime.strptime(v, fmt).strftime("%Y-%m-%d")
            except ValueError: continue
        return v

    def set(self, iso_str):
        if not iso_str: self._var.set("")
        else:
            try: self._var.set(datetime.strptime(iso_str, "%Y-%m-%d").strftime("%d/%m/%Y"))
            except ValueError: self._var.set(iso_str)

class CalPicker(tk.Frame):
    def __init__(self, master, var, top):
        super().__init__(master, bg=C["surface"], padx=10, pady=10)
        self._var = var
        self._top = top
        try: sel = datetime.strptime(var.get().strip(), "%d/%m/%Y").date()
        except Exception: sel = date.today()
        self._year, self._month = sel.year, sel.month
        self._build()

    def _build(self):
        for w in self.winfo_children(): w.destroy()
        nav = tk.Frame(self, bg=C["surface"])
        nav.pack(fill="x", pady=(0,8))
        tk.Button(nav, text="←", command=self._prev, relief="flat", bg=C["surface"], fg=C["text"], font=("Segoe UI", 10, "bold"), cursor="hand2").pack(side="left")
        tk.Label(nav, text=f"{calendar.month_name[self._month]} {self._year}".upper(), font=("Segoe UI", 9, "bold"), bg=C["surface"], fg=C["text"], width=16).pack(side="left", expand=True)
        tk.Button(nav, text="→", command=self._next, relief="flat", bg=C["surface"], fg=C["text"], font=("Segoe UI", 10, "bold"), cursor="hand2").pack(side="right")
        
        gf = tk.Frame(self, bg=C["surface"])
        gf.pack()
        today = date.today()
        for i, d in enumerate(["LU", "MA", "ME", "GI", "VE", "SA", "DO"]):
            tk.Label(gf, text=d, width=4, font=("Segoe UI", 8, "bold"), bg=C["surface"], fg=C["muted"]).grid(row=0, column=i, pady=4)
        for r, week in enumerate(calendar.monthcalendar(self._year, self._month), start=1):
            for c_idx, day in enumerate(week):
                if day == 0: tk.Label(gf, width=4, bg=C["surface"]).grid(row=r, column=c_idx)
                else:
                    d = date(self._year, self._month, day)
                    bg = C["accent"] if d == today else C["surface"]
                    fg = "white" if d == today else C["text"]
                    tk.Button(gf, text=str(day), width=4, bg=bg, fg=fg, relief="flat", cursor="hand2", font=("Segoe UI", 9),
                              command=lambda dd=d: self._pick(dd)).grid(row=r, column=c_idx, padx=1, pady=1)

    def _prev(self):
        if self._month == 1: self._month, self._year = 12, self._year - 1
        else: self._month -= 1
        self._build()

    def _next(self):
        if self._month == 12: self._month, self._year = 1, self._year + 1
        else: self._month += 1
        self._build()

    def _pick(self, d):
        self._var.set(d.strftime("%d/%m/%Y"))
        self._top.destroy()

class ImportoEntry(tk.Entry):
    def __init__(self, master, **kw):
        width = kw.pop("width", 14)
        super().__init__(master, font=("Segoe UI", 10), relief="solid", bd=1, width=width, highlightthickness=0, fg=C["text"], **kw)
        self.bind("<FocusOut>", self._format)

    def _format(self, e=None):
        v = self.get().strip().replace(" ", "")
        if not v: return
        try:
            f = float(v.replace(",", "."))
            self.delete(0, tk.END)
            self.insert(0, f"{f:.2f}".replace(".", ","))
        except ValueError: pass

    def get_float(self):
        v = self.get().strip().replace(" ", "").replace(",", ".")
        try: return float(v)
        except ValueError: return 0.0

    def set_float(self, value):
        self.delete(0, tk.END)
        if value is not None:
            self.insert(0, f"{float(value):.2f}".replace(".", ","))

# NUOVO PANNELLO IMPOSTAZIONI (Sostituisce ConfigCondizioniWindow)
class ConfigImpostazioniWindow(tk.Toplevel):
    def __init__(self, master, on_close_callback=None):
        super().__init__(master)
        self.title("Pannello Impostazioni")
        self.geometry("500x460")
        self.resizable(False, False)
        self.configure(bg=C["surface"])
        self._on_close_callback = on_close_callback
        self._build()
        self.grab_set()

    def _build(self):
        hdr = tk.Frame(self, bg=C["bg"], padx=16, pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Pannello Configurazioni Generali", font=("Segoe UI", 12, "bold"), bg=C["bg"], fg=C["text"]).pack(anchor="w")
        
        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        
        tab_cond = tk.Frame(notebook, bg=C["surface"])
        tab_metodi = tk.Frame(notebook, bg=C["surface"])
        tab_tipi = tk.Frame(notebook, bg=C["surface"])
        
        notebook.add(tab_cond, text="Condizioni Pagamento")
        notebook.add(tab_metodi, text="Metodi Pagamento")
        notebook.add(tab_tipi, text="Tipi Documento")
        
        self._build_tab(tab_cond, "config_condizioni", "Nuova stringa automatica (es. Riba 120 gg):")
        self._build_tab(tab_metodi, "config_metodi_pag", "Nuovo metodo pagamento (es. PayPal):")
        self._build_tab(tab_tipi, "config_tipi_doc", "Nuovo tipo documento (es. Preventivo):")
        
        btn_fr = tk.Frame(self, bg=C["surface"], padx=16, pady=8)
        btn_fr.pack(fill="x", side="bottom")
        tk.Button(btn_fr, text="Chiudi e Aggiorna", command=self._chiudi, bg=C["border"], fg=C["text"], font=("Segoe UI", 9, "bold"), relief="flat", padx=14, pady=4).pack(side="right")

    def _build_tab(self, parent, table_name, label_text):
        input_fr = tk.Frame(parent, bg=C["surface"], pady=12, padx=12)
        input_fr.pack(fill="x")
        
        tk.Label(input_fr, text=label_text, font=("Segoe UI", 9), bg=C["surface"], fg=C["muted"]).pack(anchor="w")
        ent = tk.Entry(input_fr, font=("Segoe UI", 10), relief="solid", bd=1)
        ent.pack(side="left", fill="x", expand=True, padx=(0, 8), pady=4, ipady=2)
        
        tree = ttk.Treeview(parent, columns=("Nome"), show="headings", height=8)
        tree.heading("Nome", text="Valori in Elenco")
        tree.column("Nome", width=300)
        
        def load_data():
            for item in tree.get_children(): tree.delete(item)
            with get_conn() as conn:
                for r in conn.execute(f"SELECT nome FROM {table_name} ORDER BY id ASC").fetchall():
                    tree.insert("", "end", values=r)
                    
        def add_data():
            val = ent.get().strip()
            if not val: return
            try:
                with get_conn() as conn: 
                    conn.execute(f"INSERT INTO {table_name} (nome) VALUES (?)", (val,))
                ent.delete(0, tk.END)
                load_data()
            except sqlite3.IntegrityError:
                messagebox.showwarning("Attenzione", "Valore già esistente.")
                
        def del_data():
            sel = tree.selection()
            if not sel: return
            val = tree.item(sel[0], "values")[0]
            with get_conn() as conn: 
                conn.execute(f"DELETE FROM {table_name} WHERE nome=?", (val,))
            load_data()

        tk.Button(input_fr, text="Aggiungi", command=add_data, bg=C["accent"], fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=12).pack(side="right", ipady=2)
        
        list_fr = tk.Frame(parent, bg=C["surface"], padx=12)
        list_fr.pack(fill="both", expand=True)
        
        tree.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(list_fr, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        
        del_fr = tk.Frame(parent, bg=C["surface"], padx=12, pady=12)
        del_fr.pack(fill="x")
        tk.Button(del_fr, text="Rimuovi Selezionata", command=del_data, bg=C["scaduta"], fg=C["scaduta_fg"], font=("Segoe UI", 9, "bold"), relief="flat", padx=10, pady=4).pack(side="left")
        
        load_data()

    def _chiudi(self):
        if self._on_close_callback: self._on_close_callback()
        self.destroy()

class AziendeWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Anagrafiche Clienti / Fornitori")
        self.geometry("720x480") 
        self.resizable(False, False)
        self.configure(bg=C["surface"])
        self._build()
        self._load_aziende()
        self.grab_set()

    def _build(self):
        hdr = tk.Frame(self, bg=C["bg"], padx=16, pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Scadenze di Default per Anagrafica", font=("Segoe UI", 12, "bold"), bg=C["bg"], fg=C["text"]).pack(anchor="w")
        
        main_fr = tk.Frame(self, bg=C["surface"], padx=16, pady=12)
        main_fr.pack(fill="both", expand=True)

        left_fr = tk.Frame(main_fr, bg=C["surface"])
        left_fr.pack(side="left", fill="both", expand=True, padx=(0, 12))

        self.tree = ttk.Treeview(left_fr, columns=("Nome", "Condizioni"), show="headings", height=12)
        self.tree.heading("Nome", text="Ragione Sociale")
        self.tree.heading("Condizioni", text="Dilazione Applicata")
        self.tree.column("Nome", width=240)
        self.tree.column("Condizioni", width=140)
        self.tree.pack(side="left", fill="both", expand=True)

        sb = ttk.Scrollbar(left_fr, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")

        right_fr = tk.Frame(main_fr, bg=C["surface"], bd=1, relief="solid", highlightthickness=0, borderwidth=1, padx=12, pady=12)
        right_fr.pack(side="right", fill="y", pady=2)

        tk.Label(right_fr, text="Scheda Anagrafica", font=("Segoe UI", 10, "bold"), bg=C["surface"], fg=C["text"]).pack(anchor="w", pady=(0,10))
        tk.Label(right_fr, text="Ragione Sociale:", font=("Segoe UI", 9), bg=C["surface"], fg=C["muted"]).pack(anchor="w")
        self.ent_nome = tk.Entry(right_fr, font=("Segoe UI", 10), relief="solid", bd=1, width=24)
        self.ent_nome.pack(anchor="w", pady=(2, 10), ipady=2)

        tk.Label(right_fr, text="Termini Pagamento:", font=("Segoe UI", 9), bg=C["surface"], fg=C["muted"]).pack(anchor="w")
        cond_values = get_condizioni_pagamento()
        self.cmb_cond = ttk.Combobox(right_fr, values=cond_values, width=22, state="readonly")
        if cond_values: self.cmb_cond.set(cond_values[0])
        self.cmb_cond.pack(anchor="w", pady=(2, 16))

        tk.Button(right_fr, text="Salva Modifiche", command=self._save_azienda, bg=C["accent"], fg="white", font=("Segoe UI", 9, "bold"), relief="flat", pady=6).pack(fill="x", pady=2)
        tk.Button(right_fr, text="Elimina", command=self._delete_azienda, bg=C["scaduta"], fg=C["scaduta_fg"], font=("Segoe UI", 9, "bold"), relief="flat", pady=6).pack(fill="x", pady=2)

        self.tree.bind("<<TreeviewSelect>>", self._on_select)

    def _load_aziende(self):
        for item in self.tree.get_children(): self.tree.delete(item)
        with get_conn() as conn:
            for r in conn.execute("SELECT nome, condizioni FROM aziende ORDER BY nome ASC").fetchall():
                self.tree.insert("", "end", values=r)

    def _on_select(self, e):
        sel = self.tree.selection()
        if not sel: return
        vals = self.tree.item(sel[0], "values")
        self.ent_nome.delete(0, tk.END)
        self.ent_nome.insert(0, vals[0])
        self.cmb_cond.set(vals[1])

    def _save_azienda(self):
        doc = self.ent_nome.get().strip()
        cond = self.cmb_cond.get()
        if not doc: return
        
        with get_conn() as conn:
            c = conn.cursor()
            c.execute("SELECT id FROM aziende WHERE nome=?", (doc,))
            if c.fetchone():
                c.execute("UPDATE aziende SET condizioni=? WHERE nome=?", (cond, doc))
            else:
                c.execute("INSERT INTO aziende (nome, condizioni) VALUES (?, ?)", (doc, cond))
                
        self.ent_nome.delete(0, tk.END)
        self._load_aziende()

    def _delete_azienda(self):
        sel = self.tree.selection()
        if not sel: return
        nome = self.tree.item(sel[0], "values")[0]
        if messagebox.askyesno("Conferma", f"Rimuovere '{nome}' dalla rubrica?"):
            with get_conn() as conn:
                conn.execute("DELETE FROM aziende WHERE nome=?", (nome,))
            self._load_aziende()

class FatturaForm(tk.Toplevel):
    def __init__(self, master, tipo="passivo", fattura=None, on_save=None):
        super().__init__(master)
        self.title("Anagrafica Documento")
        self.resizable(False, False)
        self._tipo, self._fattura, self._on_save = tipo, fattura, on_save
        self._widgets = {}
        self._build()
        if fattura: self._populate(fattura)
        self.grab_set()

    def _build(self):
        self.configure(bg=C["surface"])
        hdr = tk.Frame(self, bg=C["bg"], padx=16, pady=12)
        hdr.pack(fill="x")
        lbl = "Nuova Fattura Passiva (Fornitore)" if self._tipo == "passivo" else "Nuova Fattura Attiva (Cliente)"
        if self._fattura: lbl = f"Modifica Documento ID #{self._fattura[0]}"
        tk.Label(hdr, text=lbl, font=("Segoe UI", 11, "bold"), bg=C["bg"], fg=C["text"]).pack(anchor="w")
        
        fr = tk.Frame(self, bg=C["surface"], padx=20, pady=12)
        fr.pack()
        
        stati = STATI_PASSIVO if self._tipo == "passivo" else STATI_ATTIVO
        fields = [("Numero Fattura", "numero"), ("Ragione Sociale", "anagrafica"), ("Tipo Documento", "tipo_doc"),
                  ("Data Emissione", "data_doc"), ("Scadenza Netta", "data_scad"), ("Data Pagamento", "data_pag"),
                  ("Importo Lordo €", "importo"), ("Volume Regolato €", "pagato"), ("Stato Corrente", "stato"), 
                  ("Canale Preferito", "metodo"), ("Note / Tag", "note")]
                  
        for i, (lbl_t, key) in enumerate(fields):
            tk.Label(fr, text=lbl_t, font=("Segoe UI", 9), bg=C["surface"], fg=C["muted"]).grid(row=i, column=0, sticky="w", pady=4, padx=(0,12))
            if key in ("data_doc", "data_scad", "data_pag"): 
                w = DateEntry(fr, bg=C["surface"])
            elif key in ("importo", "pagato"): 
                w = ImportoEntry(fr)
            elif key == "stato":
                w = ttk.Combobox(fr, values=stati, width=22, state="readonly")
                w.set(stati[0])
                w.bind("<<ComboboxSelected>>", self._on_stato_changed)
            elif key == "metodo": 
                w = ttk.Combobox(fr, values=get_metodi_pagamento(), width=22) # Dinamico
            elif key == "tipo_doc":
                tipi = get_tipi_documento()
                w = ttk.Combobox(fr, values=tipi, width=22, state="readonly") # Dinamico
                if tipi: w.set(tipi[0])
            elif key == "anagrafica":
                with get_conn() as conn:
                    aziende = [r[0] for r in conn.execute("SELECT nome FROM aziende ORDER BY nome ASC").fetchall()]
                w = ttk.Combobox(fr, values=aziende, width=22)
            else: 
                w = tk.Entry(fr, font=("Segoe UI", 10), relief="solid", bd=1, width=24)
            w.grid(row=i, column=1, sticky="w", pady=4, ipady=1 if key not in ("data_doc","data_scad","data_pag","stato","metodo","tipo_doc") else 0)
            self._widgets[key] = w

        # Aggiunta Binding per Auto-allineamento "Vista fattura" / "Contanti"
        self._widgets["anagrafica"].bind("<<ComboboxSelected>>", self._check_align_vista_fattura)
        self._widgets["anagrafica"].bind("<FocusOut>", self._check_align_vista_fattura)
        self._widgets["data_doc"]._entry.bind("<FocusOut>", lambda e: [self._widgets["data_doc"]._on_focus_out(), self._check_align_vista_fattura()])

        bf = tk.Frame(self, bg=C["surface"], padx=20, pady=14)
        bf.pack(fill="x")
        tk.Button(bf, text="Salva Scheda", command=self._save, font=("Segoe UI", 9, "bold"), bg=C["accent"], fg="white", relief="flat", padx=16, pady=6).pack(side="right")
        tk.Button(bf, text="Esci", command=self.destroy, font=("Segoe UI", 9, "bold"), bg=C["border"], fg=C["text"], relief="flat", padx=16, pady=6).pack(side="right", padx=4)

    def _check_align_vista_fattura(self, event=None):
        ana = self._widgets["anagrafica"].get().strip()
        dd = self._widgets["data_doc"].get()
        if ana and dd:
            with get_conn() as conn:
                res = conn.execute("SELECT condizioni FROM aziende WHERE nome=?", (ana,)).fetchone()
                if res:
                    cond = res[0].lower()
                    if "vista fattura" in cond or "contanti" in cond:
                        self._widgets["data_scad"].set(dd)
                        self._widgets["data_pag"].set(dd)

    def _on_stato_changed(self, event=None):
        stato_sel = self._widgets["stato"].get()
        if stato_sel in ("Pagata", "Incassata"):
            totale = self._widgets["importo"].get_float()
            if totale > 0: self._widgets["pagato"].set_float(totale)
            if not self._widgets["data_pag"].get():
                self._widgets["data_pag"].set(date.today().strftime("%Y-%m-%d"))

    def _populate(self, f):
        def _set(k, v):
            w = self._widgets[k]
            if isinstance(w, ImportoEntry): w.set_float(v)
            elif isinstance(w, DateEntry): w.set(v or "")
            elif isinstance(w, ttk.Combobox): w.set(v or "")
            else: w.delete(0, tk.END); w.insert(0, v or "")
            
        _set("numero", f[2]); _set("anagrafica", f[3]); _set("data_doc", f[4]); _set("data_scad", f[5]); _set("data_pag", f[6])
        _set("importo", f[7]); _set("pagato", f[8]); _set("stato", f[9]); _set("metodo", f[10])
        _set("tipo_doc", f[11]); _set("note", f[12])

    def _save(self):
        num = self._widgets["numero"].get().strip()
        ana = self._widgets["anagrafica"].get().strip()
        if not num or not ana: return
        
        with get_conn() as conn:
            c = conn.cursor()
            vals = (self._tipo, num, ana, self._widgets["data_doc"].get(), self._widgets["data_scad"].get(), self._widgets["data_pag"].get(),
                    self._widgets["importo"].get_float(), self._widgets["pagato"].get_float(), self._widgets["stato"].get(),
                    self._widgets["metodo"].get(), self._widgets["tipo_doc"].get(), self._widgets["note"].get().strip())
            if self._fattura:
                c.execute("""UPDATE fatture SET tipo=?, numero=?, anagrafica=?, data_doc=?, data_scad=?, data_pag=?,
                           importo=?, pagato=?, stato=?, metodo=?, tipo_doc=?, note=? WHERE id=?""", vals + (self._fattura[0],))
            else:
                c.execute("""INSERT INTO fatture (tipo, numero, anagrafica, data_doc, data_scad, data_pag, importo, pagato, stato, metodo, tipo_doc, note)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""", vals)
                                  
        if self._on_save: self._on_save()
        self.destroy()

class ImportWindow(tk.Toplevel):
    def __init__(self, master, on_done=None):
        super().__init__(master)
        self.title("Modulo Importazione")
        self.geometry("450x260")
        self.resizable(False, False)
        self._on_done = on_done
        self.configure(bg=C["surface"])
        self._build()
        self.grab_set()

    def _build(self):
        hdr = tk.Frame(self, bg=C["bg"], padx=16, pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Acquisizione Tracciati Esterni", fg=C["text"], bg=C["bg"], font=("Segoe UI", 11, "bold")).pack(anchor="w")
        
        fr = tk.Frame(self, bg=C["surface"], padx=20, pady=16)
        fr.pack(fill="both", expand=True)
        
        self._path_var = tk.StringVar()
        tk.Entry(fr, textvariable=self._path_var, width=30, relief="solid", bd=1).grid(row=0, column=0, padx=2, ipady=2)
        tk.Button(fr, text="Sfoglia...", command=self._browse, bg=C["border"], fg=C["text"], font=("Segoe UI", 9, "bold"), relief="flat", padx=10).grid(row=0, column=1, padx=4)
        
        self._tipo = tk.StringVar(value="passivo")
        ttk.Combobox(fr, textvariable=self._tipo, values=["passivo (fornitori)", "attivo (clienti)"], width=24, state="readonly").grid(row=1, column=0, sticky="w", pady=12)
        
        self._log = tk.Label(fr, text="", font=("Segoe UI", 9, "bold"), bg=C["surface"], fg=C["green"])
        self._log.grid(row=2, columnspan=2, pady=4, sticky="w")
        
        tk.Button(fr, text="Esegui Importazione", command=self._import, bg=C["accent"], fg="white", font=("Segoe UI", 10, "bold"), relief="flat", pady=6).grid(row=3, columnspan=2, sticky="ew")

    def _browse(self):
        p = filedialog.askopenfilename(filetypes=[("Excel/CSV Assets", "*.xlsx *.xls *.csv")])
        if p: self._path_var.set(p)

    def _import(self):
        path = self._path_var.get().strip()
        if not path or not os.path.exists(path): return
        t = "passivo" if "passivo" in self._tipo.get() else "attivo"
        ext = os.path.splitext(path)[1].lower()
        try:
            ins, sal = self._process_xlsx(path, t) if ext in (".xlsx", ".xls") else self._process_csv(path, t)
            self._log.config(text=f"Import completato. Inseriti: {ins} | Saltati: {sal}")
            if self._on_done: self._on_done()
        except Exception as e: messagebox.showerror("Errore Import", str(e))

    def _process_xlsx(self, path, t):
        if not OPENPYXL_OK: return 0, 0
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows: return 0, 0
        hd = [str(x).strip() if x else "" for x in rows[0]]
        ins = sal = 0
        with get_conn() as conn:
            c = conn.cursor()
            for r in rows[1:]:
                rec = {hd[i]: r[i] for i in range(min(len(hd), len(r)))}
                num = str(rec.get("Numero", "") or "").strip()
                ana = str(rec.get("Fornitore", rec.get("Cliente", "")) or "").strip()
                if not num or not ana: continue
                
                c.execute("INSERT OR IGNORE INTO aziende (nome, condizioni) VALUES (?, ?)", (ana, "30 Giorni"))
                
                if c.execute("SELECT id FROM fatture WHERE numero=? AND anagrafica=? AND tipo=?", (num, ana, t)).fetchone():
                    sal += 1; continue
                dd = parse_date(rec.get("Data"))
                ds = calcola_scadenza_automatico(c, ana, dd)
                imp = parse_importo(rec.get("Totale"))
                td = str(rec.get("Tipo documento") or "Fattura").strip()
                if "differita" in td.lower(): td = "Fattura differita"
                st = mappa_stato_fte(rec.get("Stato FTE"), t)
                c.execute("INSERT INTO fatture (tipo,numero,anagrafica,data_doc,data_scad,importo,pagato,stato,tipo_doc) VALUES (?,?,?,?,?,?,?,?,?)",
                          (t, num, ana, dd, ds, imp, 0.0, st, td))
                ins += 1
        return ins, sal

    def _process_csv(self, path, t):
        ins = sal = 0
        with open(path, newline="", encoding="utf-8-sig") as f:
            rdr = csv.DictReader(f)
            with get_conn() as conn:
                c = conn.cursor()
                for rec in rdr:
                    num = str(rec.get("Numero", "") or "").strip()
                    ana = str(rec.get("Fornitore", rec.get("Cliente", "")) or "").strip()
                    if not num or not ana: continue
                    
                    c.execute("INSERT OR IGNORE INTO aziende (nome, condizioni) VALUES (?, ?)", (ana, "30 Giorni"))
                    
                    if c.execute("SELECT id FROM fatture WHERE numero=? AND anagrafica=? AND tipo=?", (num, ana, t)).fetchone():
                        sal += 1; continue
                    dd = parse_date(rec.get("Data"))
                    ds = calcola_scadenza_automatico(c, ana, dd)
                    imp = parse_importo(rec.get("Totale"))
                    td = str(rec.get("Tipo documento") or "Fattura").strip()
                    if "differita" in td.lower(): td = "Fattura differita"
                    st = mappa_stato_fte(rec.get("Stato FTE"), t)
                    c.execute("INSERT INTO fatture (tipo,numero,anagrafica,data_doc,data_scad,importo,pagato,stato,tipo_doc) VALUES (?,?,?,?,?,?,?,?,?)",
                              (t, num, ana, dd, ds, imp, 0.0, st, td))
                    ins += 1
        return ins, sal

class ExportMassivoDialog(tk.Toplevel):
    def __init__(self, master, on_confirm):
        super().__init__(master)
        self.title("Seleziona Intervallo Corrispettivi")
        self.geometry("360x220")
        self.resizable(False, False)
        self.configure(bg=C["surface"])
        self.on_confirm = on_confirm
        self._build()
        self.grab_set()

    def _build(self):
        tk.Label(self, text="Seleziona Periodo da Esportare", font=("Segoe UI", 11, "bold"), bg=C["surface"], fg=C["text"]).pack(pady=12)
        
        fr = tk.Frame(self, bg=C["surface"])
        fr.pack(pady=4)

        tk.Label(fr, text="Anno:", font=("Segoe UI", 9), bg=C["surface"]).grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.spin_anno = tk.Spinbox(fr, from_=2020, to=2035, width=8, font=("Segoe UI", 10))
        self.spin_anno.delete(0, "end")
        self.spin_anno.insert(0, str(datetime.now().year))
        self.spin_anno.grid(row=0, column=1, padx=8, pady=4)

        mesi_nomi = [f"{i:02d} - {calendar.month_name[i].capitalize()}" for i in range(1, 13)]

        tk.Label(fr, text="Dal mese:", font=("Segoe UI", 9), bg=C["surface"]).grid(row=1, column=0, sticky="w", padx=4, pady=4)
        self.cmb_da = ttk.Combobox(fr, values=mesi_nomi, width=16, state="readonly")
        self.cmb_da.set(mesi_nomi[0])
        self.cmb_da.grid(row=1, column=1, padx=8, pady=4)

        tk.Label(fr, text="Al mese:", font=("Segoe UI", 9), bg=C["surface"]).grid(row=2, column=0, sticky="w", padx=4, pady=4)
        self.cmb_a = ttk.Combobox(fr, values=mesi_nomi, width=16, state="readonly")
        self.cmb_a.set(mesi_nomi[datetime.now().month - 1])
        self.cmb_a.grid(row=2, column=1, padx=8, pady=4)

        btn_fr = tk.Frame(self, bg=C["surface"])
        btn_fr.pack(fill="x", side="bottom", pady=12, padx=16)
        
        tk.Button(btn_fr, text="Genera Report Completo", command=self._confirm, bg=C["green"], fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=12, pady=4).pack(side="right")
        tk.Button(btn_fr, text="Annulla", command=self.destroy, bg=C["border"], fg=C["text"], font=("Segoe UI", 9), relief="flat", padx=12, pady=4).pack(side="left")

    def _confirm(self):
        anno = int(self.spin_anno.get())
        m_da = int(self.cmb_da.get().split(" - ")[0])
        m_a = int(self.cmb_a.get().split(" - ")[0])
        if m_da > m_a:
            messagebox.showerror("Errore", "Il mese di inizio non può essere successivo al mese di fine.")
            return
        self.destroy()
        self.on_confirm(anno, m_da, m_a)


# ─── APPLICAZIONE CORE (VIDALOCA PREMIUM) ────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Scadenziario Cash Flow 2026 — VIDALOCA")
        self.geometry("1360x860")
        self.configure(bg=C["bg"])
        init_db()
        
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", font=("Segoe UI", 9), rowheight=28, background=C["surface"], fieldbackground=C["surface"], foreground=C["text"])
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"), background=C["border"], foreground=C["text"], relief="flat")
        style.map("Treeview", background=[("selected", C["accent"])], foreground=[("selected", "white")])
        
        self._view = "dashboard" 
        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", lambda *_: self._load_data())
        
        # Filtri Date Aggiuntivi
        self._search_da = tk.StringVar()
        self._search_da.trace_add("write", lambda *_: self._load_data())
        self._search_a = tk.StringVar()
        self._search_a.trace_add("write", lambda *_: self._load_data())

        self._sort_col, self._sort_rev = "data_scad", False
        self._current_tag_filter = "tutti"  
        
        # Variabili Corrispettivi
        self._corr_anno = tk.IntVar(value=datetime.now().year)
        self._corr_mese = tk.IntVar(value=datetime.now().month)
        self._corr_entries = {} 

        # Variabili Dashboard (Filtri in alto)
        self._dash_anno = tk.IntVar(value=datetime.now().year)
        self._dash_mese = tk.IntVar(value=datetime.now().month)
        self._dash_filtro_globale = tk.BooleanVar(value=False)

        self._build_sidebar()
        
        self.main_container = tk.Frame(self, bg=C["bg"])
        self.main_container.pack(side="left", fill="both", expand=True)
        
        self.dashboard_frame = tk.Frame(self.main_container, bg=C["bg"])
        self.table_frame = tk.Frame(self.main_container, bg=C["bg"])
        self.corrispettivi_frame = tk.Frame(self.main_container, bg=C["bg"])
        
        self._build_table_layout()
        self._build_corrispettivi_layout()
        self._set_view("dashboard")

    def _build_sidebar(self):
        sb = tk.Frame(self, bg=C["text"], width=230) 
        sb.pack(side="left", fill="y")
        sb.pack_propagate(False)
        
        brand_fr = tk.Frame(sb, bg=C["text"], pady=16)
        brand_fr.pack(fill="x")
        
        loaded_logo = False
        if PILLOW_OK:
            try:
                req = urllib.request.Request(LOGO_URL, headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req, timeout=5) as response:
                    raw_data = response.read()
                img_data = io.BytesIO(raw_data)
                img = Image.open(img_data)
                img.thumbnail((200, 110), Image.Resampling.LANCZOS)
                self._logo_img = ImageTk.PhotoImage(img)
                lbl_logo = tk.Label(brand_fr, image=self._logo_img, bg=C["text"])
                lbl_logo.pack()
                loaded_logo = True
            except Exception: pass
                
        if not loaded_logo:
            tk.Label(brand_fr, text="VIDALOCA", font=("Segoe UI", 16, "bold"), bg=C["text"], fg=C["accent"]).pack()
            tk.Label(brand_fr, text="di Michela Vidale", font=("Segoe UI", 9, "italic"), bg=C["text"], fg=C["muted"]).pack(pady=(2,0))
        
        sep_top = tk.Frame(sb, bg=C["muted"], height=1)
        sep_top.pack(fill="x", padx=16, pady=(10, 14))

        self._nav_btns = {}
        navs = [
            ("dashboard", "🏠   Dashboard"), 
            ("passivo", "📉   Registro Passivo"),
            ("attivo", "📈   Registro Attivo"), 
            ("tutte", "📋   Tutti i Documenti"),
            ("corrispettivi", "📊   Gestione Corrispettivi")
        ]
        for key, text in navs:
            b = tk.Button(sb, text=text, font=("Segoe UI", 10), bg=C["text"], fg="#dcd1c4", anchor="w",
                          relief="flat", padx=20, pady=10, cursor="hand2", activebackground=C["accent"], activeforeground="white")
            b.config(command=lambda k=key: self._set_view(k))
            b.pack(fill="x", padx=8, pady=2)
            self._nav_btns[key] = b
            
        sep = tk.Frame(sb, bg=C["muted"], height=1)
        sep.pack(fill="x", padx=16, pady=16)
        
        tk.Button(sb, text="🏢   Rubrica Aziende", font=("Segoe UI", 9), bg=C["text"], fg="#dcd1c4", anchor="w", relief="flat", padx=20, pady=6, command=lambda: AziendeWindow(self)).pack(fill="x", padx=8)
        tk.Button(sb, text="📥   Import Gestionale", font=("Segoe UI", 9), bg=C["text"], fg="#dcd1c4", anchor="w", relief="flat", padx=20, pady=6, command=lambda: ImportWindow(self, self._refresh_current_view)).pack(fill="x", padx=8)
        tk.Button(sb, text="⚙️   Impostazioni", font=("Segoe UI", 9), bg=C["text"], fg="#dcd1c4", anchor="w", relief="flat", padx=20, pady=6, command=lambda: ConfigImpostazioniWindow(self, on_close_callback=self._refresh_current_view)).pack(fill="x", padx=8)

        spacer = tk.Frame(sb, bg=C["text"])
        spacer.pack(fill="both", expand=True)

        footer_fr = tk.Frame(sb, bg=C["text"], pady=12)
        footer_fr.pack(fill="x", side="bottom")
        tk.Label(footer_fr, text="Fatto con ❤️ da ", font=("Segoe UI", 8), bg=C["text"], fg="#bfae9e").pack(side="left", padx=(16, 0))
        lbl_link = tk.Label(footer_fr, text="Giovanni Pio", font=("Segoe UI", 8, "bold", "underline"), bg=C["text"], fg=C["accent"], cursor="hand2")
        lbl_link.pack(side="left")
        lbl_link.bind("<Button-1>", lambda _: webbrowser.open("https://familiarigiovannipio.it"))

    def _build_table_layout(self):
        self._kpi_frame = tk.Frame(self.table_frame, bg=C["bg"])
        self._kpi_frame.pack(fill="x", padx=24, pady=(20, 10))
        
        tb = tk.Frame(self.table_frame, bg=C["bg"])
        tb.pack(fill="x", padx=24, pady=8)
        
        search_fr = tk.Frame(tb, bg="white", bd=1, relief="solid", highlightthickness=0)
        search_fr.pack(side="left", ipady=2)
        search_fr.config(highlightbackground=C["border"])
        
        tk.Label(search_fr, text="  🔍 Cerca Nome/Note: ", bg="white", fg=C["muted"]).pack(side="left")
        tk.Entry(search_fr, textvariable=self._search_var, font=("Segoe UI", 10), bg="white", relief="flat", bd=0, width=20).pack(side="left", padx=4)
        
        tk.Label(search_fr, text=" | Da (AAAA-MM-DD): ", bg="white", fg=C["muted"]).pack(side="left")
        tk.Entry(search_fr, textvariable=self._search_da, font=("Segoe UI", 10), bg="white", relief="flat", bd=0, width=11).pack(side="left", padx=4)
        
        tk.Label(search_fr, text=" | A (AAAA-MM-DD): ", bg="white", fg=C["muted"]).pack(side="left")
        tk.Entry(search_fr, textvariable=self._search_a, font=("Segoe UI", 10), bg="white", relief="flat", bd=0, width=11).pack(side="left", padx=4)
        
        tk.Button(tb, text="💥  Elimina", command=self._delete_invoice, bg=C["surface"], fg=C["red"], font=("Segoe UI", 9, "bold"), relief="solid", borderwidth=1, padx=14, pady=5).pack(side="right", padx=3)
        tk.Button(tb, text="✏️  Modifica", command=self._edit_invoice, bg=C["surface"], fg=C["text"], font=("Segoe UI", 9, "bold"), relief="solid", borderwidth=1, padx=14, pady=5).pack(side="right", padx=3)
        tk.Button(tb, text="➕  Nuovo Documento", command=self._new_invoice, bg=C["accent"], fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=16, pady=6).pack(side="right", padx=3)
        tk.Button(tb, text="📊  Esporta Excel", command=self._export_to_excel, bg=C["green"], fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=16, pady=6).pack(side="right", padx=12)

        f = tk.Frame(self.table_frame, bg="white", bd=1, relief="solid")
        f.pack(fill="both", expand=True, padx=24, pady=8)
        
        cols = ("id", "tipo", "numero", "anagrafica", "tipo_doc", "data_doc", "data_scad", "data_pag", "importo", "pagato", "residuo", "metodo", "stato", "note")
        self.tree = ttk.Treeview(f, columns=cols, show="headings", selectmode="browse")
        
        hd = {"id": "ID", "tipo": "Reg.", "numero": "N. Fattura", "anagrafica": "Azienda / Ragione Sociale",
              "tipo_doc": "Tipo Doc.", "data_doc": "Data Doc.", "data_scad": "Scadenza", "data_pag": "Data Pag.", "importo": "Totale Lordo",
              "pagato": "Pagato", "residuo": "Residuo Aperto", "metodo": "Canale", "stato": "Stato", "note": "Note"}
        for c, text in hd.items(): self.tree.heading(c, text=text, command=lambda _c=c: self._sort(_c))
            
        self.tree.column("id", width=40, anchor="center")
        self.tree.column("tipo", width=60, anchor="center")
        self.tree.column("numero", width=90)
        self.tree.column("anagrafica", width=200)
        self.tree.column("tipo_doc", width=90)
        self.tree.column("data_doc", width=75, anchor="center")
        self.tree.column("data_scad", width=75, anchor="center")
        self.tree.column("data_pag", width=75, anchor="center")
        self.tree.column("importo", width=85, anchor="e")
        self.tree.column("pagato", width=85, anchor="e")
        self.tree.column("residuo", width=85, anchor="e")
        self.tree.column("metodo", width=80, anchor="center")
        self.tree.column("stato", width=90, anchor="center")
        self.tree.column("note", width=120)
        self.tree.pack(side="left", fill="both", expand=True)
        
        sb = ttk.Scrollbar(f, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        
        self.tree.tag_configure("scaduta", background=C["scaduta"], foreground=C["scaduta_fg"])
        self.tree.tag_configure("urgente", background=C["urgente"], foreground=C["urgente_fg"])
        self.tree.tag_configure("ok", background=C["ok"], foreground=C["ok_fg"])
        self.tree.tag_configure("aperti", background=C["aperti"], foreground=C["aperti_fg"])
        self.tree.bind("<Double-1>", lambda _: self._edit_invoice())

        # Area dei totali misti dinamici (Sotto la tabella)
        self.tot_dyn_fr = tk.Frame(self.table_frame, bg=C["surface"], bd=1, relief="solid")
        self.tot_dyn_fr.pack(fill="x", padx=24, pady=(0, 8))
        
        tk.Label(self.tot_dyn_fr, text="📊 TOTALI DELLA VISTA CORRENTE (FILTRATI):", font=("Segoe UI", 9), bg=C["surface"], fg=C["muted"]).pack(side="left", padx=16, pady=6)
        self.lbl_dyn_lordo = tk.Label(self.tot_dyn_fr, text="Tot. Lordo: 0,00 €", font=("Segoe UI", 10, "bold"), bg=C["surface"], fg=C["text"])
        self.lbl_dyn_lordo.pack(side="left", padx=16, pady=6)
        self.lbl_dyn_pagato = tk.Label(self.tot_dyn_fr, text="Tot. Pagato: 0,00 €", font=("Segoe UI", 10, "bold"), bg=C["surface"], fg=C["green"])
        self.lbl_dyn_pagato.pack(side="left", padx=16, pady=6)
        self.lbl_dyn_residuo = tk.Label(self.tot_dyn_fr, text="Residuo Aperto: 0,00 €", font=("Segoe UI", 10, "bold"), bg=C["surface"], fg=C["red"])
        self.lbl_dyn_residuo.pack(side="left", padx=16, pady=6)

        # Legenda con Filtri Tag
        leg = tk.Frame(self.table_frame, bg=C["bg"])
        leg.pack(fill="x", padx=24, pady=(4, 10))
        tk.Label(leg, text="Filtra tabella per stato: ", font=("Segoe UI", 9, "italic"), bg=C["bg"], fg=C["muted"]).pack(side="left", padx=(0, 4))
        
        self.btn_f_scaduto = tk.Button(leg, text="  Scaduto  ", font=("Segoe UI", 8, "bold"), bg=C["scaduta"], fg=C["scaduta_fg"], bd=1, relief="solid", cursor="hand2", command=lambda: self._set_tag_filter("scaduta"))
        self.btn_f_scaduto.pack(side="left", padx=3)
        self.btn_f_urgente = tk.Button(leg, text="  In scadenza (7gg)  ", font=("Segoe UI", 8, "bold"), bg=C["urgente"], fg=C["urgente_fg"], bd=1, relief="solid", cursor="hand2", command=lambda: self._set_tag_filter("urgente"))
        self.btn_f_urgente.pack(side="left", padx=3)
        self.btn_f_aperti = tk.Button(leg, text="  Da Pagare / Aperti  ", font=("Segoe UI", 8, "bold"), bg=C["aperti"], fg=C["aperti_fg"], bd=1, relief="solid", cursor="hand2", command=lambda: self._set_tag_filter("aperti"))
        self.btn_f_aperti.pack(side="left", padx=3)
        self.btn_f_ok = tk.Button(leg, text="  Chiuso / Saldato  ", font=("Segoe UI", 8, "bold"), bg=C["ok"], fg=C["ok_fg"], bd=1, relief="solid", cursor="hand2", command=lambda: self._set_tag_filter("ok"))
        self.btn_f_ok.pack(side="left", padx=3)
        self.btn_f_tutti = tk.Button(leg, text="  ❌ Mostra Tutti  ", font=("Segoe UI", 8, "bold"), bg=C["surface"], fg=C["text"], bd=1, relief="solid", cursor="hand2", command=lambda: self._set_tag_filter("tutti"))
        self.btn_f_tutti.pack(side="left", padx=(12, 3))


    # ─── MODULO REGISTRO CORRISPETTIVI AVANZATO ─────────────────────────────────
    def _build_corrispettivi_layout(self):
        hdr_fr = tk.Frame(self.corrispettivi_frame, bg=C["bg"], padx=24, pady=10)
        hdr_fr.pack(fill="x")
        tk.Label(hdr_fr, text="REGISTRO CRONOLOGICO DEI CORRISPETTIVI", font=("Segoe UI", 13, "bold"), bg=C["bg"], fg=C["text"]).pack(anchor="w")
        tk.Label(hdr_fr, text="Compilazione intelligente: '.' convertito in ',' | Invio per scendere riga | Calcolo totali automatico.", font=("Segoe UI", 9, "italic"), bg=C["bg"], fg=C["muted"]).pack(anchor="w")

        ctrl_bar = tk.Frame(self.corrispettivi_frame, bg=C["surface"], bd=1, relief="solid", highlightthickness=0, padx=16, pady=10)
        ctrl_bar.pack(fill="x", padx=24, pady=4)
        ctrl_bar.config(highlightbackground=C["border"])

        tk.Label(ctrl_bar, text="Anno:", font=("Segoe UI", 9, "bold"), bg=C["surface"]).pack(side="left", padx=(0,4))
        spin_anno = tk.Spinbox(ctrl_bar, from_=2020, to=2035, textvariable=self._corr_anno, width=6, font=("Segoe UI", 10), command=self._load_corrispettivi_mese)
        spin_anno.pack(side="left", padx=(0,12))

        tk.Label(ctrl_bar, text="Mese:", font=("Segoe UI", 9, "bold"), bg=C["surface"]).pack(side="left", padx=(0,4))
        self.cmb_mese = ttk.Combobox(ctrl_bar, values=[f"{i:02d} - {calendar.month_name[i].capitalize()}" for i in range(1,13)], width=14, state="readonly")
        self.cmb_mese.set(f"{self._corr_mese.get():02d} - {calendar.month_name[self._corr_mese.get()].capitalize()}")
        self.cmb_mese.pack(side="left", padx=(0,12))
        self.cmb_mese.bind("<<ComboboxSelected>>", lambda e: [self._corr_mese.set(int(self.cmb_mese.get().split(" - ")[0])), self._load_corrispettivi_mese()])

        tk.Frame(ctrl_bar, bg=C["border"], width=1).pack(side="left", fill="y", padx=16)

        # Inserimento Massivo vincolato alla colonna Annotazioni / Autoconsumo
        tk.Label(ctrl_bar, text="⚡ Valore Autoconsumo Massivo €:", font=("Segoe UI", 9, "bold"), bg=C["surface"], fg=C["teal_accent"]).pack(side="left", padx=(0,6))
        self.ent_mass_val = tk.Entry(ctrl_bar, width=10, font=("Segoe UI", 9), relief="solid", bd=1)
        self.ent_mass_val.insert(0, "25,00")
        self.ent_mass_val.pack(side="left", padx=4)
        tk.Button(ctrl_bar, text="Applica a tutto il Mese", command=self._applica_autoconsumo_massivo, bg=C["teal_accent"], fg="white", font=("Segoe UI", 8, "bold"), relief="flat", padx=10).pack(side="left", padx=6)

        # Sezione dei due pulsanti Excel richiesti
        tk.Frame(ctrl_bar, bg=C["border"], width=1).pack(side="left", fill="y", padx=16)
        tk.Button(ctrl_bar, text="📊 Excel Mese", command=self._export_corrispettivi_mensile, bg=C["green"], fg="white", font=("Segoe UI", 8, "bold"), relief="flat", padx=10).pack(side="left", padx=2)
        tk.Button(ctrl_bar, text="📅 Excel Massivo...", command=lambda: ExportMassivoDialog(self, self._export_corrispettivi_massivo_intervallo), bg="#047857", fg="white", font=("Segoe UI", 8, "bold"), relief="flat", padx=10).pack(side="left", padx=2)

        # Corpo Tabellare con Canvas
        outer_canvas_frame = tk.Frame(self.corrispettivi_frame, bg="white", bd=1, relief="solid")
        outer_canvas_frame.pack(fill="both", expand=True, padx=24, pady=6)

        canvas = tk.Canvas(outer_canvas_frame, bg="white", highlightthickness=0)
        v_scroll = ttk.Scrollbar(outer_canvas_frame, orient="vertical", command=canvas.yview)
        self.scrollable_table_frame = tk.Frame(canvas, bg=C["border"])

        self.scrollable_table_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.scrollable_table_frame, anchor="nw")
        canvas.configure(yscrollcommand=v_scroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        v_scroll.pack(side="right", fill="y")

        headers = [
            ("Giorno", 55), ("Totale Giornaliero (€)", 140), 
            ("IVA Opz. 1 (€)", 105), ("IVA Opz. 2 (€)", 105), 
            ("IVA Opz. 3 (€)", 105), ("IVA Opz. 4 (€)", 105),
            ("Ventilati (€)", 105), ("Esenti (€)", 105), 
            ("Annotazioni/Autoconsumo (€)", 180), ("Note Testo", 140)
        ]
        
        for col_idx, (text, width) in enumerate(headers):
            lbl = tk.Label(self.scrollable_table_frame, text=text, font=("Segoe UI", 9, "bold"), bg=C["border"], fg=C["text"], width=width//8, height=2)
            lbl.grid(row=0, column=col_idx, padx=1, pady=1, sticky="nsew")

        self._corr_entries.clear()
        keys_totali = ["iva_1", "iva_2", "iva_3", "iva_4", "ventilati", "esenti", "autoconsumo"]

        for g in range(1, 32):
            row_bg = C["surface"] if g % 2 == 0 else C["bg"]
            
            lbl_g = tk.Label(self.scrollable_table_frame, text=f"{g:02d}", font=("Segoe UI", 9, "bold"), bg=row_bg, fg=C["text"], width=6)
            lbl_g.grid(row=g, column=0, padx=1, pady=1, sticky="nsew")
            
            row_dict = {}
            
            # 1. Casella Totale Giornaliero (Sola Lettura / Calcolata)
            ent_tot = tk.Entry(self.scrollable_table_frame, font=("Segoe UI", 9, "bold"), justify="right", relief="flat", bg=row_bg, fg=C["teal_accent"], state="readonly")
            ent_tot.grid(row=g, column=1, padx=1, pady=1, ipady=3, sticky="nsew")
            row_dict["totale"] = ent_tot
            
            # 2. Generazione Colonne di Input Numerico
            for col_idx, k in enumerate(keys_totali, start=2):
                ent = tk.Entry(self.scrollable_table_frame, font=("Segoe UI", 9), justify="right", relief="flat", bg=row_bg, fg=C["text"])
                ent.grid(row=g, column=col_idx, padx=1, pady=1, ipady=3, sticky="nsew")
                
                # Binding Tastiera Avanzati
                ent.bind("<KeyRelease>", lambda e, r=g, key=k: self._on_field_key_release(e, r, key))
                ent.bind("<Return>", lambda e, r=g: self._focus_next_row(r))
                row_dict[k] = ent
                
            # 3. Nota Libera in Coda
            ent_n = tk.Entry(self.scrollable_table_frame, font=("Segoe UI", 9), relief="flat", bg=row_bg, fg=C["text"])
            ent_n.grid(row=g, column=9, padx=1, pady=1, ipady=3, sticky="nsew")
            ent_n.bind("<Return>", lambda e, r=g: self._focus_next_row(r))
            row_dict["note_testo"] = ent_n
            
            self._corr_entries[g] = row_dict

        # Layout dei Totali Complessivi in Basso
        footer_pane = tk.Frame(self.corrispettivi_frame, bg=C["bg"], padx=24, pady=8)
        footer_pane.pack(fill="x", side="bottom")

        # Tabellina riepilogativa per i 3 macro-totali richiesti
        self.lbl_tot_netto = tk.Label(footer_pane, text="Totale Netto Corrispettivi: 0,00 €", font=("Segoe UI", 10, "bold"), bg=C["bg"], fg=C["text"])
        self.lbl_tot_netto.pack(side="left", padx=10)

        self.lbl_tot_autoconsumo = tk.Label(footer_pane, text="Totale Autoconsumo: 0,00 €", font=("Segoe UI", 10, "bold"), bg=C["bg"], fg=C["red"])
        self.lbl_tot_autoconsumo.pack(side="left", padx=10)

        self.lbl_tot_lordo = tk.Label(footer_pane, text="Totale Complessivo (Lordo): 0,00 €", font=("Segoe UI", 11, "bold"), bg=C["bg"], fg=C["teal_accent"])
        self.lbl_tot_lordo.pack(side="left", padx=20)

        tk.Button(footer_pane, text="💾  Salva Registro Mensile", command=self._save_corrispettivi_db, bg=C["accent"], fg="white", font=("Segoe UI", 10, "bold"), relief="flat", padx=20, pady=6).pack(side="right")

    def _on_field_key_release(self, event, r, key):
        """ Sostituisce in tempo reale il punto con la virgola e aggiorna il totale di riga """
        ent = self._corr_entries[r][key]
        val = ent.get()
        if "." in val:
            cursor_pos = ent.index(tk.INSERT)
            val = val.replace(".", ",")
            ent.delete(0, tk.END)
            ent.insert(0, val)
            ent.icursor(cursor_pos)
            
        self._ricalcola_totale_giornaliero(r)

    def _focus_next_row(self, current_row):
        """ Sposta il focus sul campo corrispondente del giorno successivo premendo Invio """
        next_r = current_row + 1
        if next_r in self._corr_entries:
            # Trova quale campo aveva il focus attivo e lo sposta giù
            focused_widget = self.focus_get()
            for key, widget in self._corr_entries[current_row].items():
                if widget == focused_widget:
                    target = self._corr_entries[next_r][key]
                    if target.cget("state") != "disabled":
                        target.focus_set()
                        target.selection_range(0, tk.END)
                    break

    def _ricalcola_totale_giornaliero(self, r):
        f = self._corr_entries[r]
        keys = ["iva_1", "iva_2", "iva_3", "iva_4", "ventilati", "esenti", "autoconsumo"]
        somma = 0.0
        for k in keys:
            somma += parse_importo(f[k].get())
            
        f["totale"].config(state="normal")
        f["totale"].delete(0, tk.END)
        if somma > 0:
            f["totale"].insert(0, f"{somma:.2f}".replace(".", ","))
        f["totale"].config(state="readonly")
        self._aggiorna_somme_mensili_interfaccia()

    def _aggiorna_somme_mensili_interfaccia(self):
        tot_netto = 0.0
        tot_auto = 0.0
        
        keys_netto = ["iva_1", "iva_2", "iva_3", "iva_4", "ventilati", "esenti"]
        for g, fields in self._corr_entries.items():
            if fields["iva_1"].cget("state") == "disabled": continue
            for k in keys_netto:
                tot_netto += parse_importo(fields[k].get())
            tot_auto += parse_importo(fields["autoconsumo"].get())
            
        self.lbl_tot_netto.config(text=f"Totale Netto Corrispettivi: {fmt_num(tot_netto)} €")
        self.lbl_tot_autoconsumo.config(text=f"Totale Autoconsumo: {fmt_num(tot_auto)} €")
        self.lbl_tot_lordo.config(text=f"Totale Complessivo (Lordo): {fmt_num(tot_netto + tot_auto)} €")

    def _load_corrispettivi_mese(self):
        anno = self._corr_anno.get()
        mese = self._corr_mese.get()
        
        for g, fields in self._corr_entries.items():
            for k, entry in fields.items():
                if k == "totale":
                    entry.config(state="normal")
                    entry.delete(0, tk.END)
                    entry.config(state="readonly")
                else:
                    entry.delete(0, tk.END)
                
        _, max_giorni = calendar.monthrange(anno, mese)
        for g, fields in self._corr_entries.items():
            state_val = "normal" if g <= max_giorni else "disabled"
            bg_color = (C["surface"] if g % 2 == 0 else C["bg"]) if g <= max_giorni else "#e2ded7"
            for k, entry in fields.items():
                if k == "totale":
                    entry.config(bg=bg_color, state="readonly")
                else:
                    entry.config(state=state_val, bg=bg_color)

        with get_conn() as conn:
            rows = conn.execute("SELECT giorno, totale, iva_1, iva_2, iva_3, iva_4, ventilati, esenti, autoconsumo, note FROM corrispettivi WHERE anno=? AND mese=?", (anno, mese)).fetchall()
            
        for r in rows:
            g, tot, iv1, iv2, iv3, iv4, vent, ese, auto, nt = r
            if g in self._corr_entries:
                f = self._corr_entries[g]
                if iv1 != 0: f["iva_1"].insert(0, f"{iv1:.2f}".replace(".", ","))
                if iv2 != 0: f["iva_2"].insert(0, f"{iv2:.2f}".replace(".", ","))
                if iv3 != 0: f["iva_3"].insert(0, f"{iv3:.2f}".replace(".", ","))
                if iv4 != 0: f["iva_4"].insert(0, f"{iv4:.2f}".replace(".", ","))
                if vent != 0: f["ventilati"].insert(0, f"{vent:.2f}".replace(".", ","))
                if ese != 0: f["esenti"].insert(0, f"{ese:.2f}".replace(".", ","))
                if auto != 0: f["autoconsumo"].insert(0, f"{auto:.2f}".replace(".", ","))
                if nt: f["note_testo"].insert(0, str(nt))
                
                # Forza il calcolo visivo della prima cella
                self._ricalcola_totale_giornaliero(g)
                
        self._aggiorna_somme_mensili_interfaccia()

    def _applica_autoconsumo_massivo(self):
        val_str = self.ent_mass_val.get().strip()
        if not val_str: return
        
        anno = self._corr_anno.get()
        mese = self._corr_mese.get()
        _, max_giorni = calendar.monthrange(anno, mese)

        for g in range(1, max_giorni + 1):
            f = self._corr_entries[g]
            f["autoconsumo"].delete(0, tk.END)
            f["autoconsumo"].insert(0, val_str)
            self._ricalcola_totale_giornaliero(g)
            
        messagebox.showinfo("Successo", "Importo autoconsumo massivo applicato a tutte le righe del mese.")

    def _save_corrispettivi_db(self):
        anno = self._corr_anno.get()
        mese = self._corr_mese.get()
        _, max_giorni = calendar.monthrange(anno, mese)
        
        with get_conn() as conn:
            c = conn.cursor()
            for g in range(1, max_giorni + 1):
                f = self._corr_entries[g]
                tot = parse_importo(f["totale"].get())
                v1 = parse_importo(f["iva_1"].get())
                v2 = parse_importo(f["iva_2"].get())
                v3 = parse_importo(f["iva_3"].get())
                v4 = parse_importo(f["iva_4"].get())
                vent = parse_importo(f["ventilati"].get())
                ese = parse_importo(f["esenti"].get())
                auto = parse_importo(f["autoconsumo"].get())
                nt = f["note_testo"].get().strip()
                
                c.execute("""
                    INSERT INTO corrispettivi (anno, mese, giorno, totale, iva_1, iva_2, iva_3, iva_4, ventilati, esenti, autoconsumo, note)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(anno, mese, giorno) DO UPDATE SET
                        totale=excluded.totale, iva_1=excluded.iva_1, iva_2=excluded.iva_2,
                        iva_3=excluded.iva_3, iva_4=excluded.iva_4, ventilati=excluded.ventilati,
                        esenti=excluded.esenti, autoconsumo=excluded.autoconsumo, note=excluded.note
                """, (anno, mese, g, tot, v1, v2, v3, v4, vent, ese, auto, nt))
                
        messagebox.showinfo("Salvataggio", f"Registro corrispettivi salvato correttamente per {calendar.month_name[mese]} {anno}.")
        self._load_corrispettivi_mese()

    # ─── ESPORTAZIONI STRUTTURATE EXCEL REGISTRI ──────────────────────────────
    def _export_corrispettivi_mensile(self):
        if not OPENPYXL_OK: return
        out_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not out_file: return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Corrispettivi Mensili"
            ws.views.sheetView[0].showGridLines = True
            
            # Stile Grafico Intestazioni
            fill_header = PatternFill(start_color="3D2D1F", end_color="3D2D1F", fill_type="solid")
            font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            font_bold = Font(name="Segoe UI", size=10, bold=True)
            font_regular = Font(name="Segoe UI", size=10)
            align_center = Alignment(horizontal="center", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")
            border_thin = Border(left=Side(style='thin', color='E7E1D5'), right=Side(style='thin', color='E7E1D5'),
                                 top=Side(style='thin', color='E7E1D5'), bottom=Side(style='thin', color='E7E1D5'))

            ws.append(["Data", "Totale Giornaliero", "IVA Opz. 1", "IVA Opz. 2", "IVA Opz. 3", "IVA Opz. 4", "Ventilati", "Esenti", "Annotaz./Autoconsumo", "Note"])
            for cell in ws[1]:
                cell.fill = fill_header; cell.font = font_header; cell.alignment = align_center

            anno, mese = self._corr_anno.get(), self._corr_mese.get()
            _, max_giorni = calendar.monthrange(anno, mese)
            
            for g in range(1, max_giorni+1):
                f = self._corr_entries[g]
                ws.append([
                    f"{g:02d}/{mese:02d}/{anno}",
                    parse_importo(f["totale"].get()), parse_importo(f["iva_1"].get()),
                    parse_importo(f["iva_2"].get()), parse_importo(f["iva_3"].get()),
                    parse_importo(f["iva_4"].get()), parse_importo(f["ventilati"].get()),
                    parse_importo(f["esenti"].get()), parse_importo(f["autoconsumo"].get()),
                    f["note_testo"].get()
                ])
                r_idx = ws.max_row
                ws.cell(row=r_idx, column=1).alignment = align_center
                for c in range(2, 10):
                    cell = ws.cell(row=r_idx, column=c)
                    cell.number_format = '#,##0.00 €'; cell.alignment = align_right
                for c in range(1, 11): ws.cell(row=r_idx, column=c).font = font_regular; ws.cell(row=r_idx, column=c).border = border_thin

            # Riga Totale di Chiusura Basso
            tot_row = ws.max_row + 1
            ws.cell(row=tot_row, column=1, value="TOTALE MESE").font = font_bold
            ws.cell(row=tot_row, column=1).alignment = align_center
            ws.cell(row=tot_row, column=1).border = border_thin
            
            for c in range(2, 10):
                col_letter = openpyxl.utils.get_column_letter(c)
                cell = ws.cell(row=tot_row, column=c, value=f"=SUM({col_letter}2:{col_letter}{tot_row-1})")
                cell.font = font_bold; cell.number_format = '#,##0.00 €'; cell.alignment = align_right; cell.border = border_thin
                
            ws.cell(row=tot_row, column=10).border = border_thin

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(max_len + 3, 13)

            wb.save(out_file)
            messagebox.showinfo("Export", "Report Excel Mensile salvato.")
        except Exception as e: messagebox.showerror("Errore", str(e))

    def _export_corrispettivi_massivo_intervallo(self, anno, m_da, m_a):
        if not OPENPYXL_OK: return
        out_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not out_file: return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Massivo Corrispettivi {anno}"
            ws.views.sheetView[0].showGridLines = True
            
            fill_header = PatternFill(start_color="1E6B7B", end_color="1E6B7B", fill_type="solid")
            font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            font_regular = Font(name="Segoe UI", size=10)
            align_center = Alignment(horizontal="center", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")
            border_thin = Border(left=Side(style='thin', color='E7E1D5'), right=Side(style='thin', color='E7E1D5'), top=Side(style='thin', color='E7E1D5'), bottom=Side(style='thin', color='E7E1D5'))

            ws.append(["Data", "Totale Giornaliero", "IVA Opz. 1", "IVA Opz. 2", "IVA Opz. 3", "IVA Opz. 4", "Ventilati", "Esenti", "Annotaz./Autoconsumo", "Note"])
            for cell in ws[1]: cell.fill = fill_header; cell.font = font_header; cell.alignment = align_center

            with get_conn() as conn:
                db_data = conn.execute("""SELECT mese, giorno, totale, iva_1, iva_2, iva_3, iva_4, ventilati, esenti, autoconsumo, note 
                                       FROM corrispettivi WHERE anno=? AND mese>=? AND mese<=? 
                                       ORDER BY mese ASC, giorno ASC""", (anno, m_da, m_a)).fetchall()
            
            data_map = {(r[0], r[1]): r[2:] for r in db_data}

            for m in range(m_da, m_a + 1):
                _, max_giorni = calendar.monthrange(anno, m)
                for g in range(1, max_giorni + 1):
                    row_vals = data_map.get((m, g), (0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, ""))
                    
                    ws.append([
                        f"{g:02d}/{m:02d}/{anno}",
                        row_vals[0], row_vals[1], row_vals[2], row_vals[3],
                        row_vals[4], row_vals[5], row_vals[6], row_vals[7], row_vals[8]
                    ])
                    r_idx = ws.max_row
                    ws.cell(row=r_idx, column=1).alignment = align_center
                    for c in range(2, 10):
                        cell = ws.cell(row=r_idx, column=c)
                        cell.number_format = '#,##0.00 €'; cell.alignment = align_right
                    for c in range(1, 11): 
                        ws.cell(row=r_idx, column=c).font = font_regular; ws.cell(row=r_idx, column=c).border = border_thin

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(max_len + 3, 13)

            wb.save(out_file)
            messagebox.showinfo("Export Massivo", f"Estrazione cronologica completata e salvata con successo.")
        except Exception as e: messagebox.showerror("Errore", str(e))

    def _export_to_excel(self):
        if not OPENPYXL_OK: return
        items = self.tree.get_children()
        if not items: return
        out_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Workbook", "*.xlsx")])
        if not out_file: return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = "Scadenze Flussi Cassa"
            ws.views.sheetView[0].showGridLines = True
            
            fill_header = PatternFill(start_color="3D2D1F", end_color="3D2D1F", fill_type="solid")
            font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            font_regular = Font(name="Segoe UI", size=10)
            align_center = Alignment(horizontal="center", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")
            border_thin = Border(left=Side(style='thin', color='E7E1D5'), right=Side(style='thin', color='E7E1D5'), top=Side(style='thin', color='E7E1D5'), bottom=Side(style='thin', color='E7E1D5'))

            headers = ["ID", "Registro", "N. Fattura", "Azienda / Ragione Sociale", "Tipo Documento", "Data Emissione", "Scadenza Netta", "Data Pagamento", "Importo Lordo (€)", "Volume Pagato (€)", "Residuo Aperto (€)", "Canale", "Stato Corrente", "Note"]
            ws.append(headers)
            for cell in ws[1]: cell.fill = fill_header; cell.font = font_header; cell.alignment = align_center

            for item_id in items:
                vals = self.tree.item(item_id, "values")
                imp_l = parse_importo(vals[8])
                pag_l = parse_importo(vals[9])
                res_l = parse_importo(vals[10])
                ws.append([int(vals[0]), vals[1], vals[2], vals[3], vals[4], vals[5], vals[6], vals[7], imp_l, pag_l, res_l, vals[11], vals[12], vals[13]])
                curr_row = ws.max_row
                for c_idx in range(1, 15):
                    cell = ws.cell(row=curr_row, column=c_idx)
                    cell.font = font_regular; cell.border = border_thin
                    if c_idx in (1, 2, 6, 7, 8, 12, 13): cell.alignment = align_center
                    elif c_idx in (9, 10, 11): cell.alignment = align_right; cell.number_format = '#,##0.00 €'

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(max_len + 3, 12)
            wb.save(out_file)
            messagebox.showinfo("Successo", "Esportazione scadenziario completata.")
        except Exception as e: messagebox.showerror("Errore", str(e))

    def _set_tag_filter(self, tag_name):
        self._current_tag_filter = tag_name
        for b, name in [(self.btn_f_scaduto, "scaduta"), (self.btn_f_urgente, "urgente"), (self.btn_f_aperti, "aperti"), (self.btn_f_ok, "ok"), (self.btn_f_tutti, "tutti")]:
            if name == tag_name: b.config(relief="sunken", borderwidth=2)
            else: b.config(relief="solid", borderwidth=1)
        self._load_data()

    def _set_view(self, view_name):
        self._view = view_name
        self._current_tag_filter = "tutti" 
        for k, btn in self._nav_btns.items():
            if k == view_name: btn.config(bg=C["accent"], fg="white")
            else: btn.config(bg=C["text"], fg="#dcd1c4")
            
        if view_name == "dashboard":
            self.table_frame.pack_forget()
            self.corrispettivi_frame.pack_forget()
            self.dashboard_frame.pack(fill="both", expand=True)
            self._load_dashboard()
        elif view_name == "corrispettivi":
            self.dashboard_frame.pack_forget()
            self.table_frame.pack_forget()
            self.corrispettivi_frame.pack(fill="both", expand=True)
            self._load_corrispettivi_mese()
        else:
            self.dashboard_frame.pack_forget()
            self.corrispettivi_frame.pack_forget()
            self.table_frame.pack(fill="both", expand=True)
            for b in (self.btn_f_scaduto, self.btn_f_urgente, self.btn_f_aperti, self.btn_f_ok): b.config(relief="solid", borderwidth=1)
            self.btn_f_tutti.config(relief="sunken", borderwidth=2)
            self._load_data()

    def _refresh_current_view(self):
        if self._view == "dashboard": self._load_dashboard()
        elif self._view == "corrispettivi": self._load_corrispettivi_mese()
        else: self._load_data()

    def _load_dashboard(self):
        for w in self.dashboard_frame.winfo_children(): w.destroy()
        
        # --- HEADER CON FILTRO MENSILE / GLOBALE ---
        header_fr = tk.Frame(self.dashboard_frame, bg=C["bg"], padx=24, pady=16)
        header_fr.pack(fill="x")
        
        tk.Label(header_fr, text="CRUSCOTTO DIREZIONALE FLUSSI DI CASSA — VIDALOCA", font=("Segoe UI", 14, "bold"), bg=C["bg"], fg=C["text"]).pack(side="left")
        
        filter_fr = tk.Frame(header_fr, bg=C["bg"])
        filter_fr.pack(side="right")
        
        tk.Checkbutton(filter_fr, text="🌐 Vista Globale (Ignora Filtri)", variable=self._dash_filtro_globale, command=self._load_dashboard, bg=C["bg"], activebackground=C["bg"], font=("Segoe UI", 9, "bold"), fg=C["accent_dark"]).pack(side="left", padx=(0, 16))
        
        tk.Label(filter_fr, text="Analisi Mese:", font=("Segoe UI", 9, "bold"), bg=C["bg"], fg=C["muted"]).pack(side="left", padx=(0,4))
        
        dash_spin_anno = tk.Spinbox(filter_fr, from_=2020, to=2035, textvariable=self._dash_anno, width=6, font=("Segoe UI", 10), command=self._load_dashboard)
        dash_spin_anno.pack(side="left", padx=(0,8))
        
        mesi_nomi = [f"{i:02d} - {calendar.month_name[i].capitalize()}" for i in range(1, 13)]
        dash_cmb_mese = ttk.Combobox(filter_fr, values=mesi_nomi, width=14, state="readonly")
        dash_cmb_mese.set(f"{self._dash_mese.get():02d} - {calendar.month_name[self._dash_mese.get()].capitalize()}")
        dash_cmb_mese.pack(side="left")
        dash_cmb_mese.bind("<<ComboboxSelected>>", lambda e: [self._dash_mese.set(int(dash_cmb_mese.get().split(" - ")[0])), self._load_dashboard()])

        # --- LOGICA CALCOLO DATI ---
        anno_sel = self._dash_anno.get()
        mese_sel = self._dash_mese.get()
        is_global = self._dash_filtro_globale.get()
        
        if is_global:
            dash_spin_anno.config(state="disabled")
            dash_cmb_mese.config(state="disabled")
            nome_mese_sel = "Tutto lo Storico"
        else:
            dash_spin_anno.config(state="normal")
            dash_cmb_mese.config(state="readonly")
            nome_mese_sel = calendar.month_name[mese_sel].capitalize()

        incassi_mese = 0.0
        fornitori_pagati_mese = 0.0
        rimanente_da_pagare = 0.0
        clienti_da_incassare = 0.0
        
        with get_conn() as conn:
            # 1. INCASSI NETTI (Corrispettivi)
            if is_global:
                res_corr = conn.execute("SELECT SUM(iva_1 + iva_2 + iva_3 + iva_4 + ventilati + esenti) FROM corrispettivi").fetchone()
            else:
                res_corr = conn.execute("SELECT SUM(iva_1 + iva_2 + iva_3 + iva_4 + ventilati + esenti) FROM corrispettivi WHERE anno=? AND mese=?", (anno_sel, mese_sel)).fetchone()
            incassi_mese = res_corr[0] if res_corr[0] is not None else 0.0
            
            # 2. FATTURE (Passive e Attive)
            fatture = conn.execute("SELECT tipo, importo, pagato, data_scad, data_pag, tipo_doc FROM fatture").fetchall()

        for tipo, importo, pagato, data_scad, data_pag, tipo_doc in fatture:
            if tipo_doc == "Nota di credito": continue 
            
            residuo = importo - pagato
            
            if tipo == 'passivo':
                # Pagamenti Effettuati (Fornitori Pagati) calcolati su data_pag
                if pagato > 0:
                    if is_global:
                        fornitori_pagati_mese += pagato
                    elif data_pag:
                        try:
                            d_pag = datetime.strptime(data_pag, "%Y-%m-%d")
                            if d_pag.year == anno_sel and d_pag.month == mese_sel:
                                fornitori_pagati_mese += pagato
                        except ValueError: pass
                
                # Rimanente da Pagare calcolato sulle scadenze aperte
                if residuo > 0:
                    if is_global:
                        rimanente_da_pagare += residuo
                    elif data_scad:
                        try:
                            d_scad = datetime.strptime(data_scad, "%Y-%m-%d")
                            if d_scad.year == anno_sel and d_scad.month == mese_sel:
                                rimanente_da_pagare += residuo
                        except ValueError: pass
                        
            elif tipo == 'attivo':
                # Clienti da Incassare (Entrate Attese) calcolato sulle scadenze aperte
                if residuo > 0:
                    if is_global:
                        clienti_da_incassare += residuo
                    elif data_scad:
                        try:
                            d_scad = datetime.strptime(data_scad, "%Y-%m-%d")
                            if d_scad.year == anno_sel and d_scad.month == mese_sel:
                                clienti_da_incassare += residuo
                        except ValueError: pass

        delta_cash_flow = incassi_mese - fornitori_pagati_mese

        # --- CREAZIONE DELLE CARD (2 Righe per massima completezza visiva) ---
        kpi_container_top = tk.Frame(self.dashboard_frame, bg=C["bg"], padx=20)
        kpi_container_top.pack(fill="x", pady=(6, 4))
        
        kpi_container_bottom = tk.Frame(self.dashboard_frame, bg=C["bg"], padx=20)
        kpi_container_bottom.pack(fill="x", pady=(4, 10))
        
        color_delta = C["green"] if delta_cash_flow >= 0 else C["red"]

        box_data_top = [
            (f"💰  INCASSI NETTI ({nome_mese_sel})", f"{fmt_num(incassi_mese)} €", C["teal_accent"]),
            (f"📉  FORNITORI PAGATI ({nome_mese_sel})", f"{fmt_num(fornitori_pagati_mese)} €", C["accent_dark"]),
            (f"⚖️  DELTA CASH FLOW ({nome_mese_sel})", f"{fmt_num(delta_cash_flow)} €", color_delta)
        ]
        
        box_data_bottom = [
            (f"🛒  RIMANENTE DA PAGARE ({nome_mese_sel})", f"{fmt_num(rimanente_da_pagare)} €", C["red"]),
            (f"📈  CLIENTI DA INCASSARE ({nome_mese_sel})", f"{fmt_num(clienti_da_incassare)} €", C["green"])
        ]
        
        def create_cards(container, data):
            for titolo, valore, colore in data:
                box = tk.Frame(container, bg="white", bd=1, relief="solid", borderwidth=1, padx=16, pady=14)
                box.pack(side="left", expand=True, fill="both", padx=6)
                tk.Label(box, text=titolo, font=("Segoe UI", 9, "bold"), bg="white", fg=C["muted"]).pack(anchor="w")
                tk.Label(box, text=valore, font=("Segoe UI", 15, "bold"), bg="white", fg=colore).pack(anchor="w", pady=(4,0))

        create_cards(kpi_container_top, box_data_top)
        create_cards(kpi_container_bottom, box_data_bottom)

        # --- SEZIONE TABELLA: SCADENZE IMMINENTI ---
        action_title_fr = tk.Frame(self.dashboard_frame, bg=C["bg"], padx=24)
        action_title_fr.pack(fill="x", pady=(14, 6))
        tk.Label(action_title_fr, text="ATTENZIONI RICHIESTE (Documenti già scaduti o in scadenza entro 7 Giorni)", font=("Segoe UI", 11, "bold"), bg=C["bg"], fg=C["text"]).pack(side="left")
        
        f_dash = tk.Frame(self.dashboard_frame, bg="white", bd=1, relief="solid")
        f_dash.pack(fill="both", expand=True, padx=24, pady=4)
        
        cols_dash = ("db_id", "tipo", "numero", "anagrafica", "tipo_doc", "data_scad", "residuo", "metodo", "stato", "note")
        self.tree_dash = ttk.Treeview(f_dash, columns=cols_dash, show="headings", selectmode="browse")
        for cd, txtd in {"db_id": "ID", "tipo": "Registro", "numero": "N. Documento", "anagrafica": "Ragione Sociale Azienda", "tipo_doc": "Tipo", "data_scad": "Scadenza", "residuo": "Residuo €", "metodo": "Canale", "stato": "Stato", "note": "Note"}.items(): 
            self.tree_dash.heading(cd, text=txtd)
            
        self.tree_dash.column("db_id", width=50, anchor="center")
        self.tree_dash.column("tipo", width=90, anchor="center")
        self.tree_dash.column("numero", width=120)
        self.tree_dash.column("anagrafica", width=280)
        self.tree_dash.column("tipo_doc", width=120)
        self.tree_dash.column("data_scad", width=110, anchor="center")
        self.tree_dash.column("residuo", width=110, anchor="e")
        self.tree_dash.column("metodo", width=80, anchor="center")
        self.tree_dash.column("stato", width=90, anchor="center")
        self.tree_dash.column("note", width=120)
        self.tree_dash.pack(side="left", fill="both", expand=True)
        
        sb_d = ttk.Scrollbar(f_dash, orient="vertical", command=self.tree_dash.yview)
        self.tree_dash.configure(yscrollcommand=sb_d.set); sb_d.pack(side="right", fill="y")
        
        self.tree_dash.tag_configure("scaduta", background=C["scaduta"], foreground=C["scaduta_fg"])
        self.tree_dash.tag_configure("urgente", background=C["urgente"], foreground=C["urgente_fg"])

        today_str = date.today().strftime("%Y-%m-%d")
        prox_7_giorni = (date.today() + timedelta(days=7)).strftime("%Y-%m-%d")

        with get_conn() as conn:
            rows_dash = conn.execute("""SELECT id, tipo, numero, anagrafica, tipo_doc, data_scad, importo, pagato, metodo, stato, note 
                            FROM fatture WHERE stato NOT IN ('Pagata', 'Incassata') AND data_scad <= ? ORDER BY data_scad ASC""", (prox_7_giorni,)).fetchall()
            
        for r in rows_dash:
            fid, tp, num, ana, t_doc, d_scad, imp, pag, met, st, nt = r
            res_val = imp - pag
            if res_val <= 0: continue
            tag = "scaduta" if d_scad and d_scad < today_str else "urgente"
            self.tree_dash.insert("", "end", values=(fid, tp.upper(), num, ana, t_doc, fmt_date(d_scad), fmt_num(res_val), met, st, nt), tags=(tag,))
        self.tree_dash.bind("<Double-1>", lambda _: self._edit_invoice_from_dashboard())

    def _sort(self, col):
        if self._sort_col == col: self._sort_rev = not self._sort_rev
        else: self._sort_col, self._sort_rev = col, False
        self._load_data()

    def _load_data(self):
        for item in self.tree.get_children(): self.tree.delete(item)
        query = "SELECT id, tipo, numero, anagrafica, data_doc, data_scad, data_pag, importo, pagato, stato, metodo, tipo_doc, note FROM fatture WHERE 1=1"
        params = []
        if self._view in ("passivo", "attivo"): query += " AND tipo = ?"; params.append(self._view)
        search = self._search_var.get().strip()
        if search: query += " AND (numero LIKE ? OR anagrafica LIKE ? OR note LIKE ?)"; lk = f"%{search}%"; params.extend([lk, lk, lk])
        
        # Gestione dei nuovi filtri date
        da = self._search_da.get().strip()
        a = self._search_a.get().strip()
        if da: query += " AND data_doc >= ?"; params.append(da)
        if a: query += " AND data_doc <= ?"; params.append(a)

        if self._sort_col: query += f" ORDER BY {self._sort_col} {"DESC" if self._sort_rev else "ASC"}"
            
        with get_conn() as conn: rows = conn.execute(query, params).fetchall()
            
        kpi_dovuto = kpi_scaduto = kpi_saldate = 0.0
        dyn_lordo = dyn_pagato = dyn_residuo = 0.0 # Per la striscia dei totali misti
        
        today_s, limit_u = date.today().strftime("%Y-%m-%d"), (date.today() + timedelta(days=7)).strftime("%Y-%m-%d")
        
        for r in rows:
            fid, tipo, numero, anagrafica, data_doc, data_scad, data_pag, importo, pagato, stato, metodo, tipo_doc, note = r
            residuo = max(0.0, importo - pagato)
            
            # Calcolo stato/tag visivo
            if stato in ("Pagata", "Incassata"):
                tag = "ok"
            elif data_scad and data_scad < today_s:
                tag = "scaduta"
            elif data_scad and data_scad <= limit_u:
                tag = "urgente"
            else:
                tag = "aperti"
            
            # Applica filtro status visuale dalla legenda
            if self._current_tag_filter != "tutti":
                if self._current_tag_filter == "aperti":
                    # Il filtro "aperti" mostra tutto tranne i saldati (Pagata/Incassata)
                    if stato in ("Pagata", "Incassata"): continue
                elif tag != self._current_tag_filter:
                    if self._current_tag_filter == "ok" and tag != "ok": continue
                    elif self._current_tag_filter == "scaduta" and tag != "scaduta": continue
                    elif self._current_tag_filter == "urgente" and tag != "urgente": continue

            # Se arriva qua, significa che la riga viene stampata (inserita nella tabella)
            if tipo_doc != "Nota di credito":
                kpi_dovuto += residuo
                if stato in ("Pagata", "Incassata"): kpi_saldate += 1
                elif data_scad and data_scad < today_s: kpi_scaduto += residuo

            # Somma i contatori dinamici basati esclusivamente su quello che sta per essere inserito a schermo
            dyn_lordo += importo
            dyn_pagato += pagato
            dyn_residuo += residuo

            self.tree.insert("", "end", values=(fid, tipo.upper(), numero, anagrafica, tipo_doc, fmt_date(data_doc), fmt_date(data_scad), fmt_date(data_pag), fmt_num(importo), fmt_num(pagato), fmt_num(residuo), metodo, stato, note), tags=(tag,))
        
        # Renderizza KPI originali in alto
        self._render_table_kpis(kpi_dovuto, kpi_scaduto, len(self.tree.get_children()), int(kpi_saldate))
        
        # Aggiorna la striscia dei totali dinamici in basso in tempo reale
        self.lbl_dyn_lordo.config(text=f"Tot. Lordo: {fmt_num(dyn_lordo)} €")
        self.lbl_dyn_pagato.config(text=f"Tot. Pagato: {fmt_num(dyn_pagato)} €")
        self.lbl_dyn_residuo.config(text=f"Residuo Aperto: {fmt_num(dyn_residuo)} €")

    def _render_table_kpis(self, dovuto, scaduto, totali, saldate):
        for w in self._kpi_frame.winfo_children(): w.destroy()
        lbl_txt = "Volume Aperto" if self._view == "tutte" else ("Esposizione Fornitori" if self._view == "passivo" else "Massa Incassi Attesa")
        kpis = [("Elementi Filtrati", str(totali), C["muted"]), (lbl_txt, f"{fmt_num(dovuto)} €", C["teal_accent"]), ("Di cui Scaduto Effettivo ⚠️", f"{fmt_num(scaduto)} €", C["scaduta_fg"] if scaduto > 0 else C["muted"]), ("Partite Saldate", str(saldate), C["green"])]
        for title, val, color in kpis:
            f = tk.Frame(self._kpi_frame, bg="white", bd=1, relief="solid", padx=14, pady=10)
            f.pack(side="left", expand=True, fill="x", padx=4)
            tk.Label(f, text=title, font=("Segoe UI", 8, "bold"), bg="white", fg=C["muted"]).pack(anchor="w")
            tk.Label(f, text=val, font=("Segoe UI", 13, "bold"), bg="white", fg=color).pack(anchor="w", pady=(4,0))

    def _get_selected_invoice(self, tree_widget):
        sel = tree_widget.selection()
        if not sel: return None
        fid = tree_widget.item(sel[0], "values")[0]
        with get_conn() as conn: 
            return conn.execute("SELECT id, tipo, numero, anagrafica, data_doc, data_scad, data_pag, importo, pagato, stato, metodo, tipo_doc, note FROM fatture WHERE id=?", (fid,)).fetchone()

    def _new_invoice(self): FatturaForm(self, tipo=self._view if self._view in ("passivo", "attivo") else "passivo", on_save=self._refresh_current_view)
    def _edit_invoice(self):
        f = self._get_selected_invoice(self.tree)
        if f: FatturaForm(self, tipo=f[1], fattura=f, on_save=self._refresh_current_view)
    def _edit_invoice_from_dashboard(self):
        f = self._get_selected_invoice(self.tree_dash)
        if f: FatturaForm(self, tipo=f[1], fattura=f, on_save=self._refresh_current_view)
    def _delete_invoice(self):
        f = self._get_selected_invoice(self.tree)
        if f and messagebox.askyesno("Conferma", "Eliminare la riga selezionata?"):
            with get_conn() as conn: conn.execute("DELETE FROM pagamenti WHERE fattura_id=?", (f[0],)); conn.execute("DELETE FROM fatture WHERE id=?", (f[0],))
            self._load_data()
if __name__ == "__main__":
    app = App()
    app.mainloop()
